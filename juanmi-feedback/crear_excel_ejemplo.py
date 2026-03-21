"""
Genera notas_ia.xlsx con 12 alumnos de ejemplo (versión 2).
Incluye columna Asistencia (Alta / Media / Baja).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Notas IA"

headers = ["Nombre", "Apellidos", "Nota_P1", "Nota_Final", "Asistencia", "Comentario_Profe"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Columnas: Nombre, Apellidos, Nota_P1, Nota_Final, Asistencia, Comentario_Profe
alumnos = [
    ("Laura",     "García Martínez",   8.5, 7.8, "Alta",  "Ha mejorado mucho en la segunda práctica. Muy constante."),
    ("Pablo",     "Fernández Torres",  4.0, 4.7, "Baja",  "Le cuesta el razonamiento probabilístico. Necesita repasar Bayes."),
    ("Sofía",     "López Ruiz",        9.5, 9.3, "Alta",  "Excelente alumna. Podría presentarlo a concurso."),
    ("Marcos",    "Sánchez Pérez",     6.0, 6.2, "Media", ""),
    ("Valentina", "Romero Díaz",       7.5, 7.7, "Alta",  ""),
    ("Andrés",    "Molina Castillo",   3.5, 3.7, "Baja",  "Falta de entrega en plazo. Ha penalizado bastante."),
    ("Irene",     "Jiménez Morales",   8.0, 7.8, "Media", ""),
    ("Héctor",    "Navarro Blanco",    5.5, 5.7, "Media", "Tiene base pero le falta profundidad en la memoria."),
    ("Carmen",    "Alonso Vega",       9.0, 8.8, "Alta",  ""),
    ("Diego",     "Reyes Serrano",     2.0, 2.7, "Baja",  "No ha asistido a tutorías. Recomendarle revisión con el docente."),
    ("Natalia",   "Cruz Parra",        5.0, 7.2, "Alta",  ""),  # subida notable
    ("Javier",    "Ortega Herrera",    8.0, 6.7, "Baja",  "Buen trabajo en equipo, pero la parte individual floja."),  # bajada notable
]

for row_num, (nombre, apellidos, p1, final, asistencia, comentario) in enumerate(alumnos, 2):
    ws.cell(row=row_num, column=1, value=nombre)
    ws.cell(row=row_num, column=2, value=apellidos)
    ws.cell(row=row_num, column=3, value=p1)
    ws.cell(row=row_num, column=4, value=final)
    ws.cell(row=row_num, column=5, value=asistencia)
    ws.cell(row=row_num, column=6, value=comentario)

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 55

wb.save("notas_ia.xlsx")
print("✅ Excel 'notas_ia.xlsx' actualizado con columna Asistencia.")
