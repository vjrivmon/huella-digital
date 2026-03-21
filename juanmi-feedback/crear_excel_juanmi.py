"""
Genera notas_juanmi.xlsx con 4 alumnos de ejemplo.

Estructura de columnas:
    Nombre, Apellidos,
    Nota_P1,           <- nota del parcial anterior
    Nota_Ej1, Max_Ej1, Feedback_Ej1,     <- ejercicio 1 (cálculo)
    Nota_Ej2, Max_Ej2, Feedback_Ej2,     <- pregunta 2 (teoría)
    Nota_Ej3a, Max_Ej3a, Feedback_Ej3a,  <- pregunta 3a (teoría)
    Nota_Ej3b, Max_Ej3b, Feedback_Ej3b,  <- pregunta 3b (teoría)
    Nota_P2,           <- nota de este parcial (suma de ejercicios, ya calculada)
    Media_Falta        <- media mínima que necesita en los dos parciales restantes para llegar a 4
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Notas IA"

headers = [
    "Nombre", "Apellidos",
    "Nota_P1",
    "Nota_Ej1", "Max_Ej1", "Feedback_Ej1",
    "Nota_Ej2", "Max_Ej2", "Feedback_Ej2",
    "Nota_Ej3a", "Max_Ej3a", "Feedback_Ej3a",
    "Nota_Ej3b", "Max_Ej3b", "Feedback_Ej3b",
    "Nota_P2",
    "Media_Falta",
]

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# 4 alumnos de ejemplo basados en los casos de Juanmi
# Máximos: Ej1=6.5, Ej2=1.0, Ej3a=1.5, Ej3b=1.0 → total 10.0
alumnos = [
    # Alumno 1: buena P1, baja P2, tutorías no necesarias si se espabila
    (
        "Alumno1", "Apellido1",
        8.5,                  # Nota_P1
        2.5, 6.5, "Criterio de división erróneo: divide por clase y no por atributo. El resto del ejercicio queda incorrecto.",
        0.6, 1.0, "Respuesta no del todo clara.",
        0.4, 1.5, "Respuesta poco clara, falta de profundidad.",
        0.3, 1.0, "Respuesta vaga, no llega al punto clave.",
        3.8,    # Nota_P2
        4.2,    # Media_Falta (necesita 4.2 de media en los 2 restantes para llegar a 4 global)
    ),
    # Alumno 2: P1 media, P2 baja, detecta errores (positivo)
    (
        "Alumno2", "Apellido2",
        5.5,
        2.5, 6.5, "Criterio de división incorrecto (divide por clase y no por atributo), pero detecta que hay un problema.",
        0.8, 1.0, "Casi bien.",
        0.0, 1.5, "No contesta.",
        0.0, 1.0, "No contesta.",
        3.3,
        4.7,
    ),
    # Alumno 3: P1 y P2 muy bajas, muchas carencias → sí recomendar tutorías
    (
        "Alumno3", "Apellido3",
        3.8,
        1.5, 6.5, "Formulación diferente a la vista en clase, incorrecta, con errores conceptuales y de formulación.",
        0.5, 1.0, "Detecta una forma alternativa de dividir pero no está bien justificada.",
        0.2, 1.5, "No aporta respuesta correcta.",
        0.1, 1.0, "No aporta respuesta correcta.",
        2.3,
        5.7,
    ),
    # Alumno 4: P1 baja, P2 más alta → mejora
    (
        "Alumno4", "Apellido4",
        4.2,
        4.0, 6.5, "Errores en el cálculo de una de las impurezas y en el decremento. La opción señalada como correcta no es la adecuada.",
        0.7, 1.0, "Casi bien, aunque la explicación es incompleta.",
        0.3, 1.5, "Respuesta incorrecta.",
        0.2, 1.0, "Respuesta incorrecta.",
        5.2,
        2.8,
    ),
]

for row_num, data in enumerate(alumnos, 2):
    for col_num, valor in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=valor)
        cell.alignment = Alignment(wrap_text=True)

# Anchos
anchos = [12, 18, 9, 10, 8, 55, 10, 8, 45, 11, 9, 45, 11, 9, 45, 9, 12]
for i, ancho in enumerate(anchos, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho

ws.row_dimensions[1].height = 20
for row_num in range(2, len(alumnos) + 2):
    ws.row_dimensions[row_num].height = 55

wb.save("notas_juanmi.xlsx")
print("✅ Excel 'notas_juanmi.xlsx' generado con 4 alumnos de ejemplo.")
